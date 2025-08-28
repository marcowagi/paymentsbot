"""
Reports generation service for the Telegram Finance Bot
Generates CSV and Excel reports from database tables
"""

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from models import User, Company, PaymentMethod, Request, Complaint, Ad
from config import REPORTS_DIR

logger = logging.getLogger(__name__)

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available, Excel reports will be disabled")

class ReportsService:
    def __init__(self):
        self.reports_dir = REPORTS_DIR
        self.reports_dir.mkdir(exist_ok=True, parents=True)
    
    async def generate_all_reports(self, session: AsyncSession) -> List[Path]:
        """Generate all reports and return list of file paths"""
        reports = []
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Generate CSV reports for all tables
        tables_data = await self._get_all_tables_data(session)
        
        for table_name, data in tables_data.items():
            if data:  # Only create report if there's data
                csv_file = await self._generate_csv_report(table_name, data, timestamp)
                if csv_file:
                    reports.append(csv_file)
                
                # Generate Excel report if available and data exists
                if EXCEL_AVAILABLE:
                    excel_file = await self._generate_excel_report(table_name, data, timestamp)
                    if excel_file:
                        reports.append(excel_file)
        
        logger.info(f"Generated {len(reports)} report files")
        return reports
    
    async def _get_all_tables_data(self, session: AsyncSession) -> Dict[str, List[Dict[str, Any]]]:
        """Get data from all tables"""
        tables_data = {}
        
        try:
            # Users data
            result = await session.execute(select(User))
            users = result.scalars().all()
            tables_data['users'] = [
                {
                    'id': user.id,
                    'telegram_id': user.telegram_id,
                    'customer_code': user.customer_code,
                    'name': user.name,
                    'phone': user.phone,
                    'language': user.language,
                    'is_registered': user.is_registered,
                    'created_at': user.created_at.isoformat() if user.created_at else None,
                    'updated_at': user.updated_at.isoformat() if user.updated_at else None
                }
                for user in users
            ]
            
            # Companies data
            result = await session.execute(select(Company))
            companies = result.scalars().all()
            tables_data['companies'] = [
                {
                    'id': company.id,
                    'name_ar': company.name_ar,
                    'name_en': company.name_en,
                    'is_active': company.is_active,
                    'created_at': company.created_at.isoformat() if company.created_at else None
                }
                for company in companies
            ]
            
            # Payment methods data
            result = await session.execute(select(PaymentMethod))
            payment_methods = result.scalars().all()
            tables_data['payment_methods'] = [
                {
                    'id': method.id,
                    'company_id': method.company_id,
                    'name_ar': method.name_ar,
                    'name_en': method.name_en,
                    'is_active': method.is_active,
                    'created_at': method.created_at.isoformat() if method.created_at else None
                }
                for method in payment_methods
            ]
            
            # Requests data
            result = await session.execute(select(Request))
            requests = result.scalars().all()
            tables_data['requests'] = [
                {
                    'id': request.id,
                    'user_id': request.user_id,
                    'company_id': request.company_id,
                    'payment_method_id': request.payment_method_id,
                    'request_type': request.request_type,
                    'amount': request.amount,
                    'reference': request.reference,
                    'destination_address': request.destination_address,
                    'status': request.status,
                    'admin_notes': request.admin_notes,
                    'created_at': request.created_at.isoformat() if request.created_at else None,
                    'updated_at': request.updated_at.isoformat() if request.updated_at else None
                }
                for request in requests
            ]
            
            # Complaints data
            result = await session.execute(select(Complaint))
            complaints = result.scalars().all()
            tables_data['complaints'] = [
                {
                    'id': complaint.id,
                    'user_id': complaint.user_id,
                    'subject': complaint.subject,
                    'message': complaint.message,
                    'status': complaint.status,
                    'admin_reply': complaint.admin_reply,
                    'created_at': complaint.created_at.isoformat() if complaint.created_at else None,
                    'updated_at': complaint.updated_at.isoformat() if complaint.updated_at else None
                }
                for complaint in complaints
            ]
            
            # Ads data
            result = await session.execute(select(Ad))
            ads = result.scalars().all()
            tables_data['ads'] = [
                {
                    'id': ad.id,
                    'title_ar': ad.title_ar,
                    'title_en': ad.title_en,
                    'text_ar': ad.text_ar,
                    'text_en': ad.text_en,
                    'is_active': ad.is_active,
                    'created_by': ad.created_by,
                    'created_at': ad.created_at.isoformat() if ad.created_at else None,
                    'broadcast_at': ad.broadcast_at.isoformat() if ad.broadcast_at else None
                }
                for ad in ads
            ]
            
        except Exception as e:
            logger.error(f"Error getting tables data: {e}")
        
        return tables_data
    
    async def _generate_csv_report(self, table_name: str, data: List[Dict[str, Any]], timestamp: str) -> Path:
        """Generate CSV report for a table"""
        try:
            filename = f"{table_name}_report_{timestamp}.csv"
            file_path = self.reports_dir / filename
            
            if not data:
                logger.warning(f"No data for table {table_name}")
                return None
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = data[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)
            
            logger.info(f"Generated CSV report: {filename}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error generating CSV report for {table_name}: {e}")
            return None
    
    async def _generate_excel_report(self, table_name: str, data: List[Dict[str, Any]], timestamp: str) -> Path:
        """Generate Excel report for a table"""
        if not EXCEL_AVAILABLE:
            return None
        
        try:
            filename = f"{table_name}_report_{timestamp}.xlsx"
            file_path = self.reports_dir / filename
            
            if not data:
                return None
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = table_name.title()
            
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            # Save workbook
            wb.save(file_path)
            logger.info(f"Generated Excel report: {filename}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error generating Excel report for {table_name}: {e}")
            return None
    
    def cleanup_old_reports(self, days: int = 7):
        """Clean up reports older than specified days"""
        try:
            cutoff_time = datetime.now().timestamp() - (days * 24 * 60 * 60)
            
            for file_path in self.reports_dir.glob("*"):
                if file_path.is_file() and file_path.stat().st_mtime < cutoff_time:
                    file_path.unlink()
                    logger.info(f"Deleted old report: {file_path.name}")
        except Exception as e:
            logger.error(f"Error cleaning up old reports: {e}")